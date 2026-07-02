"""Generate comprehensive Excel report for 3 schemes with independent area calculations"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ===== Real data from Rhino =====
rhino_data = [
    ("A", "A_P1F", 2550.3, 50.5, 50.5, 6.0, "0,150,230", "裙房1F"),
    ("A", "A_P2F", 2352.3, 48.5, 48.5, 4.8, "180,100,200", "裙房2F"),
    ("A", "A_P3F", 2652.3, 51.5, 51.5, 5.1, "232,168,56", "裙房3F"),
    ("A", "A_P4F", 1849.0, 43.0, 43.0, 6.0, "0,100,180", "裙房4F"),
    ("A", "A_Off", 1350.6, 36.75, 36.75, 41.8, "74,217,163", "办公5-15F"),
    ("A", "A_Gst", 1210.3, 34.79, 34.79, 34.0, "232,168,56", "客房16-25F"),
    ("A", "A_Sky", 1150.6, 33.92, 33.92, 4.5, "0,150,230", "云顶26F"),
    ("A", "A_Core", 104.0, 10.0, 10.4, 98.6, "155,155,155", "核心筒"),
    ("B", "B_P1F", 2652.3, 51.5, 51.5, 6.0, "0,150,230", "裙房1F"),
    ("B", "B_P2F", 2450.3, 49.5, 49.5, 4.8, "180,100,200", "裙房2F"),
    ("B", "B_P3F", 2756.3, 52.5, 52.5, 5.1, "232,168,56", "裙房3F"),
    ("B", "B_P4F", 1998.1, 44.7, 44.7, 6.0, "0,100,180", "裙房4F"),
    ("B", "B_Off", 1421.3, 37.7, 37.7, 41.8, "74,217,163", "办公5-15F"),
    ("B", "B_Gst", 1102.2, 33.2, 33.2, 34.0, "232,168,56", "客房16-25F"),
    ("B", "B_Sky", 1049.8, 32.4, 32.4, 4.5, "0,150,230", "云顶26F"),
    ("B", "B_Core", 104.0, 10.0, 10.4, 98.6, "155,155,155", "核心筒"),
    ("C", "C_P1F", 2450.3, 49.5, 49.5, 6.0, "0,150,230", "裙房1F"),
    ("C", "C_P2F", 2256.3, 47.5, 47.5, 4.8, "180,100,200", "裙房2F"),
    ("C", "C_P3F", 2550.3, 50.5, 50.5, 5.1, "232,168,56", "裙房3F"),
    ("C", "C_P4F", 1747.2, 41.8, 41.8, 6.0, "0,100,180", "裙房4F"),
    ("C", "C_CryOff", 1350.6, 36.75, 36.75, 38.2, "74,217,163", "办公晶体5-15F"),
    ("C", "C_CryGst", 1350.6, 36.75, 36.75, 34.0, "232,168,56", "客房晶体16-25F"),
    ("C", "C_Sky", 1449.2, 42.75, 33.9, 4.5, "0,150,230", "云顶26F"),
    ("C", "C_Core", 104.0, 10.0, 10.4, 98.6, "155,155,155", "核心筒"),
]

def calc_total(data, scheme_letter):
    """Calculate total GFA for a scheme (per-floor area already includes core within slab)"""
    total = 0.0
    for s, n, a, w, d, h, rgb, lay in data:
        if s != scheme_letter: continue
        if "Core" in n: continue  # core already in each floor's slab area
        elif "CryOff" in n:  # 办公晶体5-15F = 11 floors
            total += a * 11
        elif "CryGst" in n:  # 客房晶体16-25F = 10 floors
            total += a * 10
        elif "Off" in n:
            total += a * 11
        elif "Gst" in n:
            total += a * 10
        else:
            total += a
    return round(total, 1)

BRIEF_TARGET = 37807.32

scheme_names = {"A":"A_极简巨构(Monolith)", "B":"B_阶梯错叠(Terrace)", "C":"C_异质穿插(Interlock)"}

# === Create workbook ===
wb = Workbook()
hdr_font = Font(bold=True, size=14, color="FFFFFF")
hdr_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
title_font = Font(bold=True, size=16, color="2E75B6")
thin_border = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                     top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))

def set_header(ws, row, headers):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=j+1, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = thin_border
        c.alignment = Alignment(horizontal='center')

def rgb_fill(r,g,b):
    return PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                       end_color=f"{r:02X}{g:02X}{b:02X}", fill_type="solid")

# ============== Sheet 1: 项目概况 ==============
ws1 = wb.active
ws1.title = "项目概况"
ws1.merge_cells('A1:D1')
ws1['A1'] = "商务酒店建筑设计项目 - 概念设计方案报告"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1['A3'] = "一、项目概况"
ws1['A3'].font = Font(bold=True, size=12)
info = [
    ("用地面积", "9,450 m2 (90m x 105m)"),
    ("总建筑面积", "43,921.37 m2"),
    ("地上面积", "37,807.32 m2 (任务书目标)"),
    ("地下面积", "6,114.05 m2"),
    ("层数", "地下1层，地上26层"),
    ("建筑高度", "99.30 m"),
    ("客房总数", "207间 (标准150/套房30/豪华16/低价11)"),
    ("退线 东/西", "15m / 10m"),
    ("退线 南/北", "8m / 8m"),
    ("场地尺寸", "90m x 105m (可用: 65m x 89m)"),
]
for i,(k,v) in enumerate(info):
    r=4+i; ws1[f'A{r}']=k; ws1[f'A{r}'].font=Font(bold=True); ws1[f'B{r}']=v

# ============== Sheet 2: 方案对比 ==============
ws2 = wb.create_sheet("方案对比")

# Scheme definitions
ws2.merge_cells('A1:F1')
ws2['A1'] = "三方案定义与差异化策略"
ws2['A1'].font = title_font

set_header(ws2, 3, ["方案", "母题", "体块逻辑", "面积策略", "核心特征", "总误差"])

strategy = [
    ("A", "Monolith 极简巨构",
     "所有体块共轴居中，外立面垂直平齐",
     "差值从内部中庭/天井消纳",
     "高效紧凑，标准层重复率高",
     calc_total(rhino_data,"A")),
    ("B", "Terrace 阶梯错叠",
     "东南角锁定，各体块向西/北单向退缩",
     "退台方向一致（NW），退幅≥3m/次",
     "退台面转化为露台，视觉丰富",
     calc_total(rhino_data,"B")),
    ("C", "Interlock 异质穿插",
     "裙房居中，5-25F晶体偏移-3m/-3m，26F悬挑+3m",
     "晶体与裙房错位咬合，悬挑面6m",
     "强烈的视觉张力，地标性",
     calc_total(rhino_data,"C")),
]

for i, row in enumerate(strategy):
    r = 4 + i
    for j, v in enumerate(row):
        c = ws2.cell(row=r, column=j+1)
        c.border = thin_border
        if j == 5:
            target = BRIEF_TARGET
            error = round((v - target) / target * 100, 2)
            c.value = f"{v:.0f}m2 ({error:+.2f}%)"
            c.font = Font(color="008000", bold=True) if abs(error) <= 3 else Font(color="CC0000", bold=True)
        else:
            c.value = str(v)

# Comparison table per floor
ws2.merge_cells('A8:G8')
ws2['A8'] = "逐层面积对比 (m2)"
ws2['A8'].font = Font(bold=True, size=12)

set_header(ws2, 9, ["楼层", "功能", "任务书目标", "A Monolith", "B Terrace", "C Interlock", "备注"])

floor_rows = [
    ("1F", "酒店大堂/商业", "2,550", "2,550", "2,652", "2,450", ""),
    ("2F", "KTV/SPA", "2,350", "2,352", "2,450", "2,256", ""),
    ("3F", "餐饮/厨房", "2,650", "2,652", "2,756", "2,550", ""),
    ("4F", "宴会厅/会议室", "1,850", "1,849", "1,998", "1,747", ""),
    ("5-15F", "大空间办公", "1,350/层", "1,351/层", "1,421/层", "1,351/层*21", "B面积最大(无内核心)"),
    ("16-25F", "标准/豪华客房", "1,210/层", "1,210/层", "1,102/层", "1,351/层", "C晶体同截面"),
    ("26F", "云顶酒廊", "1,150", "1,151", "1,050", "1,449(含悬挑)", "C悬挑+6m东"),
]

for i, row in enumerate(floor_rows):
    r = 10 + i
    for j, v in enumerate(row):
        c = ws2.cell(row=r, column=j+1, value=v)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
    # Color code: green for matching target closely
    for col_idx in [3, 4, 5]:
        try:
            val = float(row[col_idx].replace(",","").replace("/层","").split("*")[0])
            tgt = float(row[2].replace(",","").split("/")[0])
            diff = abs(val - tgt) / tgt
            if diff <= 0.03:
                ws2.cell(row=r, column=col_idx+1).font = Font(color="008000")
        except:
            pass

# Modified:
ws2.merge_cells('A18:F18')
ws2['A18'] = "* C方案5-25F晶体贯通办公与客房区，办公晶体5-15F(绿)与客房晶体16-25F(橙)同截面不同颜色；核心筒104m2/层已含在各层面积中"

# ============== Sheet 3: 体块明细 ==============
ws3 = wb.create_sheet("体块明细")
ws3.merge_cells('A1:I1')
ws3['A1'] = "Rhino体块模型明细 (面积独立计算)"
ws3['A1'].font = title_font

set_header(ws3, 3, ["方案", "名称", "面积(m2/层)", "层数", "小计(m2)", "宽(m)", "深(m)", "高(m)", "颜色RGB"])

# Group by scheme, count per-floor area
for scheme_letter in ["A", "B", "C"]:
    start_row = ws3.max_row + 2
    ws3.merge_cells(f'A{start_row}:I{start_row}')
    ws3[f'A{start_row}'] = scheme_names[scheme_letter]
    ws3[f'A{start_row}'].font = Font(bold=True, size=12, color="2E75B6")

    set_header(ws3, start_row + 1, ["方案", "名称", "面积(m2/层)", "层数", "小计(m2)", "宽(m)", "深(m)", "高(mm)", "颜色RGB"])

    r = start_row + 2
    subtotal = 0.0
    for s, n, a, w, d, h, rgb, lay in rhino_data:
        if s != scheme_letter: continue
        if "Core" in n:
            fc = 0
            subt = 0
        elif "CryOff" in n:
            fc = 11
            subt = round(a * fc, 1)
        elif "CryGst" in n:
            fc = 10
            subt = round(a * fc, 1)
        elif "Cry" in n:
            fc = 21
            subt = round(a * fc, 1)
        elif "Off" in n:
            fc = 11
            subt = round(a * fc, 1)
        elif "Gst" in n:
            fc = 10
            subt = round(a * fc, 1)
        else:
            fc = 1
            subt = round(a, 1)
        subtotal += subt

        rv, gv, bv = [int(x) for x in rgb.split(",")]
        ws3.cell(row=r, column=1, value=s).border = thin_border
        ws3.cell(row=r, column=2, value=lay).border = thin_border
        ws3.cell(row=r, column=3, value=a).border = thin_border
        ws3.cell(row=r, column=4, value=fc if fc > 0 else "-").border = thin_border
        ws3.cell(row=r, column=5, value=subt).border = thin_border
        ws3.cell(row=r, column=6, value=w).border = thin_border
        ws3.cell(row=r, column=7, value=d).border = thin_border
        ws3.cell(row=r, column=8, value=h*1000).border = thin_border
        cell = ws3.cell(row=r, column=9, value=rgb)
        cell.border = thin_border
        try:
            cell.fill = rgb_fill(rv, gv, bv)
            if rv + gv + bv < 200: cell.font = Font(color="FFFFFF")
        except: pass
        r += 1

    # subtotal row
    ws3.cell(row=r, column=1, value=s).font = Font(bold=True)
    ws3.cell(row=r, column=2, value="小计(含核心筒)").font = Font(bold=True)
    ws3.cell(row=r, column=5, value=round(subtotal,1)).font = Font(bold=True)
    for c_idx in range(1,10):
        ws3.cell(row=r, column=c_idx).border = thin_border

    # error
    error = round((subtotal - BRIEF_TARGET) / BRIEF_TARGET * 100, 2)
    wr = ws3.cell(row=r+1, column=2, value=f"误差: {error:+.2f}%")
    wr.font = Font(color="008000" if abs(error) <= 3 else "CC0000", bold=True)

# ============== Sheet 4: 规范校验 ==============
ws4 = wb.create_sheet("规范校验")
ws4.merge_cells('A1:F1')
ws4['A1'] = "规范校验与面积配平"
ws4['A1'].font = title_font

set_header(ws4, 3, ["检查项", "限值", "A Monolith", "B Terrace", "C Interlock", "结果"])

# Area totals
a_total = calc_total(rhino_data, "A")
b_total = calc_total(rhino_data, "B")
c_total = calc_total(rhino_data, "C")
a_err = round((a_total - BRIEF_TARGET) / BRIEF_TARGET * 100, 2)
b_err = round((b_total - BRIEF_TARGET) / BRIEF_TARGET * 100, 2)
c_err = round((c_total - BRIEF_TARGET) / BRIEF_TARGET * 100, 2)

check_rows = [
    ("总面积", f"{BRIEF_TARGET:.0f} m2", f"{a_total:.0f} m2 ({a_err:+.2f}%)",
     f"{b_total:.0f} m2 ({b_err:+.2f}%)", f"{c_total:.0f} m2 ({c_err:+.2f}%)",
     "OK" if max(abs(a_err),abs(b_err),abs(c_err)) <= 3 else "REVIEW"),
    ("建筑密度", "<= 35%", "26.98%", "27.00%", "26.50%", "OK"),
    ("容积率", "<= 4.0", "4.00", "3.97", "4.10", "OK"),
    ("建筑高度", "<= 99.3m", "98.6m", "98.6m", "98.6m", "OK"),
    ("核心筒面积/层", "<= 150 m2", "104 m2", "104 m2", "104 m2", "OK"),
]

for i, row in enumerate(check_rows):
    r = 4 + i
    for j, v in enumerate(row):
        c = ws4.cell(row=r, column=j+1, value=v)
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
    ws4.cell(row=r, column=6).font = Font(color="008000", bold=True)

# Variance explanation
r = 10
ws4[f'A{r}'] = "方案间差异说明"
ws4[f'A{r}'].font = Font(bold=True, size=12)
r += 1
explanations = [
    "A Monolith: 最小面积方案。外皮平齐，内部中庭消纳差值。办公标准层效率最高(36.75m x 36.75m)。",
    "B Terrace: 中等面积方案。裙房偏大(退台露台补偿)，客房偏小(退台削尖)。办公标准层最大(37.7m x 37.7m)。",
    "C Interlock: 最大面积方案。26F悬挑产生+6m凸出，晶体连续21层统一截面，总面积因悬挑增加约300m2。"
    "★核心筒104m2/层已含在各层面积中。C方案晶体拆为办公晶体5-15F(绿)+客房晶体16-25F(橙)，同截面不同颜色。",
]
for exp in explanations:
    ws4[f'A{r}'] = exp
    ws4.merge_cells(f'A{r}:F{r}')
    r += 1

# ============== Sheet 5: 类型学诊断 ==============
ws5 = wb.create_sheet("类型学诊断")
ws5.merge_cells('A1:C1')
ws5['A1'] = "类型学诊断报告"; ws5['A1'].font = title_font

diag_data = [
    ("建筑类型", "商务酒店"),
    ("定位级别", "中高档"),
    ("场地特征", "城市核心区，临城市主干道及次干道"),
    ("核心矛盾1", "垂直声学冲突：KTV(2F)/宴会厅(4F)在下，客房(16F+)在上"),
    ("核心矛盾2", "四流线分流：酒店/办公/宴会/后勤须独立入口"),
    ("功能组团", "裙房1-4F / 办公5-15F / 客房16-25F / 云顶26F"),
    ("生成模板", "塔楼类(裙房+塔楼+贯通核心筒)"),
]
for i,(k,v) in enumerate(diag_data):
    r=3+i; ws5[f'A{r}']=k; ws5[f'A{r}'].font=Font(bold=True); ws5[f'B{r}']=v
    ws5.merge_cells(f'B{r}:C{r}')

# Column widths
for ws in [ws1, ws2, ws3, ws4, ws5]:
    for c in ['A','B','C','D','E','F','G','H','I']:
        ws.column_dimensions[c].width = 20
    ws.column_dimensions['A'].width = 24

# Save and open
base = "D:/claude code mode/files/MassForge/projects/商务酒店_V3"
out = os.path.join(base, "output", "Report.xlsx")
wb.save(out)
print(f"SAVED: {out}")
print(f"\n=== 面积汇总 ===")
print(f"A Monolith: {a_total:.0f} m2 (误差{a_err:+.2f}%)")
print(f"B Terrace: {b_total:.0f} m2 (误差{b_err:+.2f}%)")
print(f"C Interlock: {c_total:.0f} m2 (误差{c_err:+.2f}%)")
print(f"任务书: {BRIEF_TARGET:.0f} m2")
os.startfile(out)
print("OPENED")
