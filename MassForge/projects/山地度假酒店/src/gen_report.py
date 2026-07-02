"""山地野奢度假酒店 — 三方案面积报告"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = Workbook()

title_font = Font(bold=True, size=14)
header_font = Font(bold=True, size=11)
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, row, labels):
    for i, l in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=l)
        c.font = header_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# ===== Sheet 1: 面积汇总 =====
ws1 = wb.active
ws1.title = "面积汇总"
ws1.merge_cells('A1:H1')
ws1['A1'] = "山地野奢度假酒店 — 三方案面积对比"
ws1['A1'].font = title_font

set_header(ws1, 3, ["功能", "目标(㎡)", "A 沿山展开", "B 围合院落", "C 线性台地", "A基底(㎡)", "B基底(㎡)", "C基底(㎡)"])

# 数据: (名称, 目标GFA, A_GFA, B_GFA, C_GFA, A_base, B_base, C_base)
data = [
    ("公建B1_后勤", 500, 500, 500, 500, 500, 500, 500),
    ("公建1F_温泉SPA", 2500, 2500, 2500, 2500, 2500, 2500, 2500),
    ("公建2F_大堂会议", 2000, 2025, 2025, 2025, 2025, 2025, 2025),
    ("公建3F_餐饮", 1500, 1520, 1520, 1520, 1520, 1520, 1520),
    ("侧翼_会议", 0, 195, 195, 195, 195, 195, 195),
    ("独栋客房(15间)", 1460, 1460, 1460, 1460, 730, 730, 730),
    ("双拼客房(20间)", 1200, 1200, 1200, 1200, 600, 600, 600),
    ("联排客房(45间)", 2250, 2250, 2250, 2250, 1125, 1125, 1125),
]

for r, d in enumerate(data, 4):
    for c, v in enumerate(d, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        cell.border = border

# 合计行
sum_row = len(data) + 4
ws1.cell(row=sum_row, column=1, value="合计").font = Font(bold=True)
gfa_sum = [sum(d[i] for d in data) for i in range(1,4)]
base_sum = [sum(d[i] for d in data) for i in range(4,7)]
for ci, val in [(2, gfa_sum[0]), (3, gfa_sum[0]), (4, gfa_sum[1]), (5, gfa_sum[2]),
                (6, base_sum[0]), (7, base_sum[1]), (8, base_sum[2])]:
    c = ws1.cell(row=sum_row, column=ci, value=val)
    c.font = Font(bold=True); c.border = border

err_row = sum_row + 1
ws1.cell(row=err_row, column=1, value="误差%").font = Font(bold=True)
for ci, val in [(2, 0), (3, round((gfa_sum[0]-12000)/12000*100,1)),
                (4, round((gfa_sum[1]-12000)/12000*100,1)),
                (5, round((gfa_sum[2]-12000)/12000*100,1))]:
    ws1.cell(row=err_row, column=ci, value=f"{val:+.1f}%").font = Font(bold=True)

ws1.cell(row=err_row+1, column=1, value="目标: 12,000㎡ | 密度≤15% = 基底≤4,700㎡ | 限高12m | 客房≥80间 | 图层 R1/A/B/功能")
for col in ['A','B','C','D','E','F','G','H']:
    ws1.column_dimensions[col].width = 18

# ===== Sheet 2: 方案说明 =====
ws2 = wb.create_sheet("方案说明")
ws2.merge_cells('A1:B1')
ws2['A1'] = "三方案布局说明"
ws2['A1'].font = title_font

desc = [
    ("方案A 沿山展开", "offset=500000"),
    ("布局", "公建在最下(南)，联排中段，双拼上段，独栋最上(北)"),
    ("逻辑", "从入口到山顶逐级而上，功能从公共到私密"),
    ("", ""),
    ("方案B 围合院落", "offset=780000"),
    ("布局", "公建在最上，联排在左侧，双拼在右侧，独栋在最下"),
    ("逻辑", "公建位于最高处观景，客房组团围合分布"),
    ("", ""),
    ("方案C 线性台地", "offset=1060000"),
    ("布局", "独栋最下，联排中段，双拼上段，公建最上"),
    ("逻辑", "私密到公共逐级上升，与方案A形成倒置对比"),
    ("", ""),
    ("图层规范", "R1/{方案字母}/{功能} (如 R1/A/独栋客房)"),
    ("色码", "公建B1=深蓝(0,100,180) 公建1F=紫(180,100,200)"),
    ("", "公建2F=蓝(0,150,230) 公建3F=橙(232,168,56)"),
    ("", "侧翼=绿(74,217,163) 独栋=棕(160,120,80)"),
    ("", "双拼=浅棕(200,170,130) 联排=米色(220,195,160)"),
    ("高度", "公建B1=3m(地下) 公建1F=5m 公建2F/3F/侧翼=4.5m"),
    ("", "独栋总统/豪华=7m(2F) 独栋标准/双拼/联排=6m(2F)"),
    ("", "最高点7m < 限高12m"),
]
for r, (k, v) in enumerate(desc, 3):
    if k: ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=v)
ws2.column_dimensions['A'].width = 24
ws2.column_dimensions['B'].width = 50

# Save
base = "D:/claude code mode/files/MassForge/projects/山地度假酒店"
out = os.path.join(base, "output", "Report.xlsx")
wb.save(out)
print(f"SAVED: {out}")
os.startfile(out)
